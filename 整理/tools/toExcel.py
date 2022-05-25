import numpy as np
from openpyxl import Workbook
def toExcel(arrayStr, yKey, xKey, excelStore):
	wb = Workbook()
	ws = wb.active
	ws.append(["NUMBER", "CONTENT"])
	for i in range(len(arrayStr)):
		if(i == xKey):
			continue
		for j in range(len(arrayStr[0])):
			if(j == yKey or len(arrayStr[i][j]) == 0):
				continue

			if(len(arrayStr[xKey][j]) == 0 and len(arrayStr[i][yKey]) == 0):
				ws.append(["--", arrayStr[i][j]])
			elif(len(arrayStr[xKey][j]) == 0):
				ws.append(["-"+arrayStr[i][yKey], arrayStr[i][j]])
			elif(len(arrayStr[i][yKey]) == 0):
				ws.append([arrayStr[xKey][j]+arrayStr[i][yKey], arrayStr[i][j]])
			else:
				ws.append([arrayStr[xKey][j]+arrayStr[i][yKey], arrayStr[i][j]])
	wb.save(excelStore)
def toExcel2(arrayStr, excelStore):
	wb = Workbook()
	ws = wb.active
	for i, line in enumerate(arrayStr):
		for j, value in enumerate(line):
			ws.cell(row=i+1, column=j+1).value = value
	# print("./"+excelStore)
	wb.save(excelStore)