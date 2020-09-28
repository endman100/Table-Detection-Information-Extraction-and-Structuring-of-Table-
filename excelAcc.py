import json
import os
import cv2
import numpy as np
from sklearn.metrics import confusion_matrix
from openpyxl import Workbook
from openpyxl import load_workbook
import csv

import torch
def loadXlsxData(path):
	wb = load_workbook(path, read_only=True)
	ws = wb.active
	count = 1
	data = [[ws.cell(row=count, column=1).value, ws.cell(row=count, column=2).value]]
	while(data[-1][0]):
		print("{}".format(count), end="\r")
		count+=1
		data.append([ws.cell(row=count, column=1).value, ws.cell(row=count, column=2).value])
	del(data[-1])
	return data[1:]

resultFlooder = "./result/marge/excel/"
GTFlooders    = "./groundtruth/excel2/"

with open('result.csv', 'w', newline='') as csvfile:
	writer = csv.writer(csvfile)

	for GTFlooderName in os.listdir(GTFlooders):
		
		GTPath = os.path.join(GTFlooders, GTFlooderName)
		resultFlooderName, ext = GTFlooderName.split("FPK_")
		resultPath = os.path.join(resultFlooder, "FPK_"+ext)
		print(resultPath, GTPath)
		# continue
		
		GTData = loadXlsxData(GTPath)
		preData = loadXlsxData(resultPath)
		# print(GTData)
		# print(preData)

		preDict = {}
		for key, value in preData:
			if(key in preDict):
				preDict[key] = None
				print("key repeat {}".format(key))
			else:
				preDict[key] = value

		GTDict = {}
		for key, value in GTData:
			if(key in GTDict):
				# GTDict[key] = None
				print("something error")
			else:
				GTDict[key] = value

		# print(preDict)
		# print(GTDict)

		count = 0
		for key in GTDict:
			if(key in preDict):
				if(preDict[key] == GTDict[key]):
					count+=1
				else:
					print("str not same key:{} --> {} <--> {}".format(key, preDict[key], GTDict[key]))
			else:
				print("key not found {} --> {}".format(key, GTDict[key]))
		Acc = count/len(GTData)
		print(Acc)
		writer.writerow([resultPath, Acc])